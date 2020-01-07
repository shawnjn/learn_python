'''
解决问题：需要将班组发出的excel数据进行汇总，发出的excel表格格式是统一;
需要操作的表格放值python程序同文件下，才能进行操作，后续可以增加选择文件位置;
打开需要编辑表格【默认为sheet(0)】;
'''

from openpyxl import load_workbook      #导入openpyxl进行excel表格操作，load_workbook 打开excel用
from openpyxl import workbook           #创建excel用


wb1 = load_workbook('sample.xlsx')
wb2 = load_workbook('sample2.xlsx')

'''激活需要编辑表格'''
ws1 = wb1.active
ws2 = wb2.active

'''主程序'''
a = 10
b = []
i = 0           #从第1行开始
j = 8           #计划从第8列（即h列）
while i < a:
    i += 1
    b.insert(i, ws2.cell(i, j).value)

'''判断ws1表内表格是否有数值，如果没有则将ws2内指定单元格内数字（内容）赋予ws1指定的单元格内'''
i = 0
while i < a:
    i += 1

    if ws1.cell(i, j).value == None:
        ws1.cell(i, j).value = b[i-1]
    else:
        continue

wb1.save('汇总表.xlsx')       #保存为新的excel表格
